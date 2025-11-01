"""
Export utilities for the attendance system
"""
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from attendance_system.utils.calculations import generate_report


def export_report(db, filename, report_data=None):
    """
    Export a comprehensive attendance report to Excel
    
    Args:
        db: DBManager instance
        filename: Path to save the Excel file
        report_data: Optional pre-generated report data
    """
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # Generate report if not provided
    if report_data is None:
        report_data = generate_report(db)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Define styles
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Student ID", "Name", "Course", "Total Sessions", "Present", "Absent", "Attendance %"]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add report data
    for student in report_data:
        row_data = [
            student.get('student_id', ''),
            student.get('name', ''),
            student.get('course', ''),
            student.get('total_sessions', 0),
            student.get('present', 0),
            student.get('absent', 0),
            f"{student.get('attendance_percentage', 0):.1f}%"
        ]
        ws.append(row_data)
    
    # Style data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Center align for numeric columns
            if cell.column in [4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.append(["Total Students", len(report_data)])
    
    if report_data:
        avg_attendance = sum(s.get('attendance_percentage', 0) for s in report_data) / len(report_data)
        summary_ws.append(["Average Attendance", f"{avg_attendance:.1f}%"])
    
    # Style summary
    for cell in summary_ws["A"]:
        cell.font = Font(bold=True)
    
    # Save the workbook
    wb.save(filename)
    return True


def export_students(db, filename):
    """
    Export students and their attendance data to Excel
    
    Args:
        db: DBManager instance
        filename: Path to save the Excel file (should end with .xlsx)
    """
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Define styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Student ID", "Name", "Course", "Email", "Total Sessions", "Present", "Attendance %"]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Get students with attendance data
    students = db.get_students_with_attendance()
    
    # Add student data
    for student in students:
        student_id, name, course, email, total_sessions, present_count = student
        
        # Calculate attendance percentage
        attendance_pct = 0
        if total_sessions > 0:
            attendance_pct = (present_count / total_sessions) * 100
        
        row_data = [
            student_id,
            name,
            course or "",
            email or "",
            total_sessions,
            present_count,
            f"{attendance_pct:.1f}%"
        ]
        ws.append(row_data)
    
    # Style data cells and auto-adjust column widths
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Center align for numeric columns
            if cell.column in [5, 6, 7]:  # Total Sessions, Present, Attendance %
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add metadata sheet
    metadata_ws = wb.create_sheet("Export Info")
    metadata_ws.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    metadata_ws.append(["Total Students", len(students)])
    metadata_ws.append(["Generated By", "Attendance System"])
    
    # Style metadata
    for cell in metadata_ws["A"]:
        cell.font = Font(bold=True)
    
    # Save the workbook
    wb.save(filename)
    return True


def export_grades(db, filename):
    """
    Export all grades to Excel
    
    Args:
        db: DBManager instance
        filename: Path to save the Excel file
    """
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"
    
    # Define styles
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Student ID", "Name", "Assessment Type", "Assessment Name", "Score", "Max Score", "Percentage", "Date"]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Get all grades
    grades = db.get_all_grades()
    
    # Add grade data
    for grade in grades:
        student_id, name, assessment_type, assessment_name, score, max_score, date = grade
        
        # Calculate percentage
        percentage = (score / max_score * 100) if max_score > 0 else 0
        
        row_data = [
            student_id,
            name,
            assessment_type,
            assessment_name,
            score,
            max_score,
            f"{percentage:.1f}%",
            date
        ]
        ws.append(row_data)
    
    # Style data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Center align for numeric columns
            if cell.column in [5, 6, 7]:  # Score, Max Score, Percentage
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)
    return True


def export_attendance_detailed(db, filename):
    """
    Export detailed attendance records to Excel
    
    Args:
        db: DBManager instance
        filename: Path to save the Excel file
    """
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Records"
    
    # Define styles
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    present_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    absent_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Date", "Student ID", "Student Name", "Course", "Status"]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Get all attendance records (you'll need to add this method to db_manager if it doesn't exist)
    try:
        # Get all students and their attendance
        students = db.get_all_students()
        all_records = []
        
        for student_id, name, course, email in students:
            # Get attendance for this student
            db.cursor.execute("""
                SELECT date, status FROM attendance 
                WHERE student_id = ? 
                ORDER BY date DESC
            """, (student_id,))
            attendance_records = db.cursor.fetchall()
            
            for date, status in attendance_records:
                all_records.append((date, student_id, name, course or "", status))
        
        # Sort by date (most recent first)
        all_records.sort(key=lambda x: x[0], reverse=True)
        
        # Add attendance data
        for record in all_records:
            ws.append(record)
        
    except Exception as e:
        print(f"Error fetching attendance records: {e}")
        return False
    
    # Style data cells
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)), start=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Color code based on status
        status_cell = ws.cell(row=row_num, column=5)
        if status_cell.value == "Present":
            status_cell.fill = present_fill
            status_cell.font = Font(bold=True, color="155724")
        elif status_cell.value == "Absent":
            status_cell.fill = absent_fill
            status_cell.font = Font(bold=True, color="721C24")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)
    return True